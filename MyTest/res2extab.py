from openpyxl import *
from openpyxl.styles import *

ft = Font(name='TimesNewRoman',size=11)
alg = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)

try:
    source_file = open('results/test_001.fzp')
    text = source_file.read().split(':').pop()
except Exception as e:
    print("Failed: %s" % e)
else:
    result_book = Workbook()
    result_sheet = result_book.get_sheet_by_name(u'Sheet')
    result_sheet.title = u'Result'
    result_columns = text.split('\n')
    for i in range(1, len(result_columns)):
        rows = result_columns[i-1].split(';')
        for j in range(1, len(rows)):
            result_sheet.cell(row=i, column=j).value = rows[j-1]
            result_sheet.cell(row=i, column=j).font = ft
            result_sheet.cell(row=i, column=j).alignment = alg
    source_file.close()
result_book.save('test_result.xlsx')