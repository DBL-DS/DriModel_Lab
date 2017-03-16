#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import *
from openpyxl.styles import *
from tkinter.messagebox import *
from tkinter.filedialog import *
from tkinter import *


# definitions
ft = Font(name='TimesNewRoman',size=11)
alg = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=True, indent=0)


# codes
class Parser:
    def __init__(self):
        self.main_dlg = Tk()
        self.main_dlg.geometry('420x120')
        self.main_dlg.resizable(False, False)
        self.main_dlg.title('XLParser')
        self.main_dlg.iconbitmap('icon\\XLParser.ico')

        Label(self.main_dlg, text='txt-xlsx转换工具', font=('微软雅黑', 24, 'bold')).grid(row=0, column=0, columnspan=4)

        Label(self.main_dlg, text='文件输入:', font=('微软雅黑', 10)).grid(row=1, column=0, sticky=W)
        self.path = Entry(self.main_dlg, width=40)
        self.path.grid(row=1, column=1, sticky=W)
        Label(self.main_dlg, text='    ').grid(row=1, column=2)
        Button(self.main_dlg, text='浏览...', command=self.__read__).grid(row=1, column=3, sticky=W)

        Button(self.main_dlg, text='开始转换', font=('微软雅黑', 11, 'bold', 'italic'), command=self.__trans__)\
                .grid(row=3, columnspan=4)
        self.main_dlg.mainloop()

    def __read__(self):
        filename = askopenfilename(initialdir=os.getcwd(), title='读取', filetypes=[('文件类型', '*.txt;*.fzp;*.att')])
        self.path.delete(0, END)
        self.path.insert(0, filename)
        self.result_name = filename.split('_').pop()

    def __trans__(self):
        try:
            source_file = open(self.path.get())
            text = source_file.read().split('$').pop()
            result_name = text.split(':').pop(0)
            text = text[len(result_name)+1:]
            if os.path.exists(r'parse_result.xlsx'):
                result_book = load_workbook('parse_result.xlsx')
                if result_book.get_sheet_names().__contains__(result_name):
                    choice = askyesno(title='数据确认', message=result_name+'已存在，是否覆盖？')
                    if choice == TRUE:
                        result_book.remove_sheet(result_book.get_sheet_by_name(result_name))
                    else:
                        return
                result_sheet = result_book.create_sheet(result_name)
            else:
                result_book = Workbook()
                result_sheet = result_book.get_sheet_by_name(u'Sheet')
                result_sheet.title = result_name
            result_columns = text.split('\n')
            for i in range(1, len(result_columns)):
                rows = result_columns[i - 1].split(';')
                for j in range(1, len(rows)):
                    result_sheet.cell(row=i, column=j).value = rows[j - 1]
                    result_sheet.cell(row=i, column=j).font = ft
                    result_sheet.cell(row=i, column=j).alignment = alg
            source_file.close()
            result_book.save('parse_result.xlsx')
        except Exception as e:
            showerror(title='错误', message=e)
        else:
            choice = askyesno(title='成功', message='转换已成功！是否继续进行转换')
            if choice == FALSE:
                self.main_dlg.quit()
                self.main_dlg.destroy()


if __name__ == '__main__':
    Parser()
